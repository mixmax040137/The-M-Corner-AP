"""สร้างไฟล์ .xlsx 11 แท็บ สำหรับอัปโหลดขึ้น Google Drive แล้วแปลงเป็น Google Sheet"""
import json, datetime, base64, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
seed = json.load(open(os.path.join(HERE, 'seed.json'), encoding='utf-8'))

wb = Workbook()
wb.remove(wb.active)

HEAD_FILL = PatternFill('solid', fgColor='1F2A44')
HEAD_FONT = Font(bold=True, color='FFFFFF', name='Tahoma', size=10)

WIDTHS = {'date': 13, 'money': 15, 'number': 10, 'multiline': 42, 'files': 34, 'text': 18, 'select': 18, 'bool': 10}

for name in seed['order']:
    cols = seed['schema'][name]
    rows = seed['data'][name]
    ws = wb.create_sheet(title=name)

    ws.append([c['label'] for c in cols])
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(c['type'], 18)
    ws.freeze_panes = 'A2'

    for r in rows:
        line = []
        for c in cols:
            v = r.get(c['key'])
            t = c['type']
            if v in (None, ''):
                line.append(None)
            elif t == 'date':
                try:
                    line.append(datetime.datetime.strptime(str(v)[:10], '%Y-%m-%d'))
                except ValueError:
                    line.append(str(v))
            elif t in ('number', 'money'):
                line.append(float(v) if v not in (None, '') else None)
            elif t == 'files':
                line.append('\n'.join(v) if isinstance(v, list) else str(v))
            else:
                line.append(str(v))
        ws.append(line)

    # รูปแบบตัวเลข/วันที่ และ dropdown
    n = len(rows) + 1
    for i, c in enumerate(cols, start=1):
        L = get_column_letter(i)
        if c['type'] == 'date':
            for row in range(2, n + 1):
                ws.cell(row=row, column=i).number_format = 'yyyy-mm-dd'
        elif c['type'] == 'money':
            for row in range(2, n + 1):
                ws.cell(row=row, column=i).number_format = '#,##0.00'
        elif c['type'] == 'number':
            for row in range(2, n + 1):
                ws.cell(row=row, column=i).number_format = '#,##0'
        elif c['type'] in ('multiline', 'files'):
            for row in range(2, n + 1):
                ws.cell(row=row, column=i).alignment = Alignment(wrap_text=True, vertical='top')

out = os.path.join(HERE, 'the-m-corner-ap-db.xlsx')
wb.save(out)
size = os.path.getsize(out)
print('sheets:', len(wb.sheetnames), '|', ', '.join(wb.sheetnames))
print('bytes:', size)
with open(out, 'rb') as f:
    b64 = base64.b64encode(f.read()).decode()
open(os.path.join(HERE, 'db.b64'), 'w').write(b64)
print('base64 chars:', len(b64))
